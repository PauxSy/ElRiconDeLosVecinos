from decimal import Decimal
from functools import wraps
import json
from django.conf import settings
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.shortcuts import get_object_or_404, render #PRUEBA PARA VER SI CONECTA CON LA BD
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.hashers import make_password
from django.shortcuts import render, redirect
import mercadopago
from .models import Bodeguero, Producto , Usuario,Administrador, Vendedor 
from django.core.mail import send_mail
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.contrib import messages
from django.urls import reverse
from django.utils.encoding import force_bytes, force_str
from django.template.loader import render_to_string
from datetime import datetime, timedelta
import os
import hashlib
from django.utils.timezone import now, timedelta
from django.core.mail import EmailMultiAlternatives
from .forms import ProductoForm, UsuarioForm
from django.views.decorators.http import require_POST
from django.utils.html import strip_tags

from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Producto 
from .forms import ProductoForm  # Asegúrate de tener un ModelForm para el modelo Producto
import openpyxl

from rincondelosvecinos import models

def crear_preferencia(request):
    if request.method == 'POST':
        try:
            # Leer y cargar el JSON desde el cuerpo de la solicitud
            body = json.loads(request.body.decode('utf-8'))  # Decodificar y cargar JSON

            print("Datos recibidos:", body)  # Log para depuración

            # Configura el SDK de Mercado Pago
            sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)

            # Calcular el total del carrito usando las claves correctas
            total = sum(item['quantity'] * item['price'] for item in body)

            print("Total calculado:", total)

            # Crear la preferencia de pago
            preference_data = {
                "items": [
                    {
                        "id": item['id'],
                        "title": item['name'],
                        "quantity": item['quantity'],
                        "unit_price": float(item['price']),
                        "currency_id": "CLP"
                    }
                    for item in body
                ],
                "back_urls": {
                    "success": "https://tusitio.com/success",
                    "failure": "https://tusitio.com/failure",
                    "pending": "https://tusitio.com/pending"
                },
                "auto_return": "approved"
            }

            # Crear la preferencia con el SDK
            preference_response = sdk.preference().create(preference_data)
            preference_id = preference_response["response"]["id"]

            print("Preference ID:", preference_id)  # Log para verificar

            # Devolver el ID de la preferencia al frontend
            return JsonResponse({"preference_id": preference_id})

        except Exception as e:
            print(f"Error al crear la preferencia: {e}")
            return JsonResponse({"error": "Error al crear la preferencia"}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)

def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Verificar si la sesión identifica al usuario como administrador
        if request.session.get('admin_autenticado', False) and request.session.get('is_admin', False):
            return view_func(request, *args, **kwargs)
        else:
            # Agregar un mensaje de error y redirigir
            messages.error(request, "No tienes permiso para acceder a esta página.")
            return redirect('vista_inicio')  # Cambia 'vista_inicio' a la vista adecuada
    return _wrapped_view

def role_required(allowed_roles):
    """
    Decorador para verificar si el usuario pertenece a uno de los roles permitidos.
    """
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            # Obtener roles del usuario desde la sesión
            is_admin = request.session.get('is_admin', False)
            is_bodeguero = request.session.get('is_bodeguero', False)
            is_vendedor = request.session.get('is_vendedor', False)
            
            # Verificar si el usuario tiene al menos uno de los roles permitidos
            if ('admin' in allowed_roles and is_admin) or \
               ('bodeguero' in allowed_roles and is_bodeguero) or \
               ('vendedor' in allowed_roles and is_vendedor):
                return view_func(request, *args, **kwargs)
            
            # Si no tiene permisos, redirigir y mostrar mensaje
            messages.error(request, "No tienes permisos para acceder a esta sección.")
            return redirect('login')  # Cambiar por la vista adecuada para redirigir
            
        return _wrapped_view
    return decorator


@admin_required
def generar_informe_inventario(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventario"

    # Agregar un título con la fecha actual
    fecha_actual = datetime.now().strftime("%d-%m-%Y")
    titulo = f"Informe Inventario - {fecha_actual}"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    sheet.cell(row=1, column=1).value = titulo
    sheet.cell(row=1, column=1).font = openpyxl.styles.Font(size=14, bold=True)
    sheet.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")

    # Agregar encabezados en la fila 3
    headers = ['ID', 'Nombre', 'Descripción', 'Stock', 'Precio', 'IVA', 'Precio Total', 'Categoría', 'Estado', 'Imagen URL']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=3, column=col_num, value=header)
        sheet.cell(row=3, column=col_num).font = openpyxl.styles.Font(bold=True)

    # Agregar datos a partir de la fila 4
    productos = Producto.objects.all()
    for row_num, producto in enumerate(productos, start=4):
        sheet.cell(row=row_num, column=1, value=producto.id)
        sheet.cell(row=row_num, column=2, value=producto.nombre)
        sheet.cell(row=row_num, column=3, value=producto.descripcion)
        sheet.cell(row=row_num, column=4, value=producto.stock)
        sheet.cell(row=row_num, column=5, value=producto.precio)
        sheet.cell(row=row_num, column=6, value=producto.iva)
        sheet.cell(row=row_num, column=7, value=producto.precio_total)
        sheet.cell(row=row_num, column=8, value=producto.categoria)
        sheet.cell(row=row_num, column=9, value=producto.estado)
        sheet.cell(row=row_num, column=10, value=producto.img_url)

    # Configurar el nombre del archivo con la fecha actual
    nombre_archivo = f"Informe_Inventario_{fecha_actual}.xlsx"

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response

@admin_required
def cambiar_imagen_producto(request, id):
    if request.method == 'POST':
        producto = get_object_or_404(Producto, id=id)
        nueva_url = request.POST.get('img_url')

        if nueva_url:
            producto.img_url = nueva_url
            producto.save()
            messages.success(request, "La imagen del producto ha sido actualizada correctamente.")
        else:
            messages.error(request, "No se pudo actualizar la imagen. Por favor, ingresa una URL válida.")

    return redirect('inventario')

def guardar_productos(request):
    if request.method == 'POST':
        errores = []
        for producto in Producto.objects.all():
            try:
                # Extraer datos específicos del producto por su ID
                img_url = request.POST.get(f'img_url_{producto.id}', producto.img_url)
                nombre = request.POST.get(f'nombre_{producto.id}', producto.nombre)
                descripcion = request.POST.get(f'descripcion_{producto.id}', producto.descripcion)
                stock = int(request.POST.get(f'stock_{producto.id}', producto.stock) or 0)
                precio = float(request.POST.get(f'precio_{producto.id}', producto.precio) or 0)
                iva = precio * 0.19  # Calcular el IVA (19%)
                precio_total = precio + iva  # Calcular el precio total
                categoria = request.POST.get(f'categoria_{producto.id}', producto.categoria)
                estado = request.POST.get(f'estado_{producto.id}', producto.estado)

                # Actualizar el producto
                producto.img_url = img_url
                producto.nombre = nombre
                producto.descripcion = descripcion
                producto.stock = stock
                producto.precio = precio
                producto.iva = round(iva, 2)  # Redondear a 2 decimales
                producto.precio_total = round(precio_total, 2)  # Redondear a 2 decimales
                producto.categoria = categoria
                producto.estado = estado

                # Guardar en la base de datos
                producto.save()

            except Exception as e:
                errores.append(f"Error en producto ID {producto.id}: {str(e)}")

        if errores:
            messages.error(request, f"Errores detectados: {', '.join(errores)}")
        else:
            messages.success(request, "Productos actualizados correctamente.")

    return redirect('inventario')

def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Verificar si el usuario está autenticado y es administrador
        if request.user.is_authenticated and getattr(request.user, 'is_admin', False):
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
    return _wrapped_view

# @admin_required
def vista_agregarproductoadmin(request):
    if not request.session.get('admin_autenticado'):  # Verifica si está autenticado
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return render(request, 'catalogo.html')

    admin_id = request.session.get('admin_id')
    if not admin_id:
        messages.error(request, 'No se puede identificar al administrador. Por favor, inicia sesión.')
        return render(request, 'inicioSesionUser.html')

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.iva = round(producto.precio * Decimal('0.19'))
            producto.admin_id = admin_id
            producto.save()
            messages.success(request, 'Producto agregado exitosamente.')  # Mensaje para SweetAlert
            return render(request, 'verinventario.html')
        else:
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = ProductoForm()

    categorias = Producto.CATEGORIAS
    return render(request, 'agregarProductoAdmin.html', {'form': form, 'categorias': categorias})



def vista_iniciouser(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        # Buscar en Usuarios, Administradores, Vendedores y Bodegueros
        usuario = Usuario.objects.filter(email=email).first()
        admin = Administrador.objects.filter(email=email).first()
        vendedor = Vendedor.objects.filter(email=email).first()
        bodeguero = Bodeguero.objects.filter(email=email).first()

        if usuario:
            # Verificar si la cuenta está bloqueada
            if usuario.bloqueado:
                messages.error(request, "Tu cuenta está bloqueada. Por favor, restablece tu contraseña para desbloquearla.")
                return render(request, 'recuperarcontrasena.html')

            # Verificar si la cuenta no está activa (no confirmada)
            if not usuario.is_active:
                if usuario.fecha_token and now() - usuario.fecha_token > timedelta(hours=23):
                    # Reenviar correo de confirmación si han pasado más de 23 horas
                    uid = urlsafe_base64_encode(force_bytes(usuario.id))
                    token = default_token_generator.make_token(usuario)
                    activation_link = request.build_absolute_uri(
                        reverse('activar_cuenta', kwargs={'uidb64': uid, 'token': token})
                    )

                    # Enviar correo de confirmación
                    email_subject = "Confirma tu cuenta en El Rincón de los Vecinos"
                    email_body = render_to_string('email_confirmacion.html', {
                        'nombre_usuario': usuario.nombre,
                        'activation_link': activation_link,
                    })
                    email = EmailMultiAlternatives(
                        subject=email_subject,
                        body=email_body,
                        from_email="tuemail@gmail.com",
                        to=[usuario.email],
                    )
                    email.attach_alternative(email_body, "text/html")
                    email.send()

                    # Actualizar la fecha del token en la base de datos
                    usuario.fecha_token = now()
                    usuario.save()

                    messages.info(request, "Te hemos enviado un nuevo correo de confirmación. Revisa tu bandeja de entrada.")
                else:
                    # Recordar que debe confirmar su cuenta si han pasado menos de 23 horas
                    messages.warning(request, "Debes confirmar tu cuenta. Revisa el correo que te enviamos para activarla.")
                return render(request, 'inicioSesionUser.html')

            # Verificar la contraseña
            if usuario.contrasena == encriptar_con_salt(password, usuario.salt):
                # Reiniciar intentos fallidos
                usuario.intentos_fallidos = 0
                usuario.bloqueado = False
                usuario.last_login = now()  # Actualizar el último inicio de sesión
                usuario.save()

                # Configurar la sesión
                request.session['usuario_autenticado'] = True
                request.session['nombre_usuario'] = usuario.nombre
                request.session['primer_apellido'] = usuario.primer_apellido
                request.session['email'] = usuario.email

                messages.success(request, "Inicio de sesión exitoso")
                return redirect('catalogo')  # Redirige al catálogo o panel del usuario
            else:
                # Incrementar intentos fallidos y bloquear si es necesario
                usuario.intentos_fallidos += 1
                if usuario.intentos_fallidos >= 3:
                    usuario.bloqueado = True
                usuario.save()
                messages.error(request, "Contraseña incorrecta")
                return render(request, 'inicioSesionUser.html')

        elif admin:
            # Verificar la contraseña para el administrador
            if admin.contrasena == encriptar_con_salt(password, admin.salt):
                # Configurar la sesión
                request.session['admin_autenticado'] = True
                request.session['email'] = admin.email
                request.session['is_admin'] = True
                request.session['admin_id'] = admin.id

                messages.success(request, "Inicio de sesión exitoso como administrador")
                return redirect('dashboard')  # Redirige al panel del administrador
            else:
                messages.error(request, "Contraseña incorrecta para administrador")
                return render(request, 'inicioSesionUser.html')

        # elif vendedor:
        #     # Verificar la contraseña para el vendedor
        #     if vendedor.contrasena == encriptar_con_salt(password, vendedor.salt):
        #         # Configurar la sesión
        #         request.session['vendedor_autenticado'] = True
        #         request.session['email'] = vendedor.email
        #         request.session['is_vendedor'] = True
        #         request.session['vendedor_id'] = vendedor.id

        #         messages.success(request, "Inicio de sesión exitoso como vendedor")
        #         return redirect('panelvendedor')  # Redirige al panel del vendedor
        #     else:
        #         messages.error(request, "Contraseña incorrecta para vendedor")
        #         return render(request, 'inicioSesionUser.html')

        # elif bodeguero:
        #     # Verificar la contraseña para el bodeguero
        #     if bodeguero.contrasena == encriptar_con_salt(password, bodeguero.salt):
        #         # Configurar la sesión
        #         request.session['bodeguero_autenticado'] = True
        #         request.session['email'] = bodeguero.email
        #         request.session['is_bodeguero'] = True
        #         request.session['bodeguero_id'] = bodeguero.id

        #         messages.success(request, "Inicio de sesión exitoso como bodeguero")
        #         return redirect('panelbodeguero')  # Redirige al panel del bodeguero
        #     else:
        #         messages.error(request, "Contraseña incorrecta para bodeguero")
        #         return render(request, 'inicioSesionUser.html')

        elif vendedor:
            # Verificar si el vendedor está activo
            if not vendedor.estado:
                messages.error(request, "El vendedor está inactivo. Contacte al administrador.")
                return render(request, 'inicioSesionUser.html')

            # Verificar la contraseña para el vendedor
            if vendedor.contrasena == encriptar_con_salt(password, vendedor.salt):
                # Configurar la sesión
                request.session['vendedor_autenticado'] = True
                request.session['email'] = vendedor.email
                request.session['is_vendedor'] = True
                request.session['vendedor_id'] = vendedor.id

                messages.success(request, "Inicio de sesión exitoso como vendedor")
                return redirect('panelvendedor')  # Redirige al panel del vendedor
            else:
                messages.error(request, "Contraseña incorrecta para vendedor")
                return render(request, 'inicioSesionUser.html')

        elif bodeguero:
            # Verificar si el bodeguero está activo
            if not bodeguero.estado:
                messages.error(request, "El bodeguero está inactivo. Contacte al administrador.")
                return render(request, 'inicioSesionUser.html')

            # Verificar la contraseña para el bodeguero
            if bodeguero.contrasena == encriptar_con_salt(password, bodeguero.salt):
                # Configurar la sesión
                request.session['bodeguero_autenticado'] = True
                request.session['email'] = bodeguero.email
                request.session['is_bodeguero'] = True
                request.session['bodeguero_id'] = bodeguero.id

                messages.success(request, "Inicio de sesión exitoso como bodeguero")
                return redirect('panelbodeguero')  # Redirige al panel del bodeguero
            else:
                messages.error(request, "Contraseña incorrecta para bodeguero")
                return render(request, 'inicioSesionUser.html')

        else:
            messages.error(request, "El email no está registrado")
            return render(request, 'inicioSesionUser.html')

    return render(request, 'inicioSesionUser.html')





def tu_vista_admin(request):
    admin_menu_items = [
        {"label": "Ver Perfil", "url_name": "perfiluser"},
        {"label": "Inventario", "url_name": "historialuser"},
        {"label": "Personal", "url_name": "historialuser"},
        {"label": "Promociones", "url_name": "historialuser"},
    ]
    return render(request, "tu_plantilla.html", {"admin_menu_items": admin_menu_items})

def has_perm(self, perm, obj=None):
        """Permitir permisos básicos."""
        return True

def has_module_perms(self, app_label):
        """Permitir acceso a los módulos de la app."""
        return True

@property
def is_staff(self):
        """Definir si el usuario es parte del staff."""
        return False


def vista_recuperarcontrasena(request):
    # Verificar si el usuario ya está autenticado
    if request.session.get("nombre_usuario") and request.session.get("primer_apellido"):
        messages.error(request, "Ya tienes una sesión iniciada. Por favor, cierra sesión antes de recuperar la contraseña.")
        return render(request, 'catalogo.html')

    if request.method == "POST":
        email = request.POST.get("email")
        if not email:
            messages.error(request, "Por favor, ingresa un correo electrónico.")
            return render(request, "recuperarcontrasena.html")

        # Buscar el usuario, administrador, vendedor o bodeguero con el correo proporcionado
        usuario = Usuario.objects.filter(email=email).first()
        administrador = Administrador.objects.filter(email=email).first()
        vendedor = Vendedor.objects.filter(email=email).first()  # Buscar vendedor
        bodeguero = Bodeguero.objects.filter(email=email).first()  # Buscar bodeguero

        if not usuario and not administrador and not vendedor and not bodeguero:
            messages.error(request, "No existe una cuenta asociada a este correo.")
            return render(request, "recuperarcontrasena.html")

        # Verificar si el vendedor o bodeguero está inactivo
        if vendedor and not vendedor.estado:
            messages.error(request, "El vendedor está inactivo. No se puede enviar el correo de recuperación.")
            return render(request, "recuperarcontrasena.html")

        if bodeguero and not bodeguero.estado:
            messages.error(request, "El bodeguero está inactivo. No se puede enviar el correo de recuperación.")
            return render(request, "recuperarcontrasena.html")

        # Generar un enlace de restablecimiento para usuario, administrador, vendedor o bodeguero
        if usuario:
            user_type = "usuario"
            user_id = usuario.id
            nombre_usuario = usuario.nombre
        elif administrador:
            user_type = "admin"
            user_id = administrador.id
            nombre_usuario = administrador.email
        elif vendedor:
            user_type = "vendedor"
            user_id = vendedor.id
            nombre_usuario = vendedor.email  # Usar el correo del vendedor
        elif bodeguero:
            user_type = "bodeguero"
            user_id = bodeguero.id
            nombre_usuario = bodeguero.email  # Usar el correo del bodeguero

        reset_link = request.build_absolute_uri(
            reverse("reset_password", args=[f"{user_type}_{user_id}"])
        )

        # Renderizar la plantilla HTML para el correo
        context = {
            "nombre_usuario": nombre_usuario,
            "reset_password_link": reset_link,
        }

        email_html_content = render_to_string("recuperarpass.html", context)
        email_text_content = strip_tags(email_html_content)  # Convertir a texto plano para clientes que no soportan HTML

        # Enviar el correo de recuperación
        try:
            send_mail(
                subject="Recuperación de contraseña - El Rincón de los Vecinos",
                message=email_text_content,
                from_email="elrincondelosvecinoschile@gmail.com",  # Cambia por tu correo
                recipient_list=[email],
                fail_silently=False,
                html_message=email_html_content,  # Mensaje en formato HTML
            )
            messages.success(request, "El correo de recuperación fue enviado exitosamente. Revisa tu bandeja de entrada.")
        except Exception as e:
            messages.error(request, f"Hubo un error al enviar el correo: {str(e)}")

    return render(request, "recuperarcontrasena.html")




def reset_password(request, user_id):
    if request.method == "POST":
        nueva_contrasena = request.POST.get("nueva_contrasena")
        confirmar_contrasena = request.POST.get("confirmar_contrasena")

        # Usar una variable temporal para procesar el ID del usuario
        temp_user_id = user_id
        if temp_user_id.startswith("usuario_"):
            temp_user_id = temp_user_id.replace("usuario_", "")
            usuario_obj = Usuario.objects.filter(id=temp_user_id).first()
        elif temp_user_id.startswith("admin_"):
            temp_user_id = temp_user_id.replace("admin_", "")
            usuario_obj = Administrador.objects.filter(id=temp_user_id).first()


        elif temp_user_id.startswith("vendedor_"):
            temp_user_id = temp_user_id.replace("vendedor_", "")
            usuario_obj = Vendedor.objects.filter(id=temp_user_id).first()  # Se añadió esta parte para vendedores


        elif temp_user_id.startswith("bodeguero_"):
            temp_user_id = temp_user_id.replace("bodeguero_", "")
            usuario_obj = Bodeguero.objects.filter(id=temp_user_id).first()  # Y aquí para bodegueros
        else:
            usuario_obj = None


        # Validar si existe el usuario o administrador correspondiente
        if not usuario_obj:
            messages.error(request, "El usuario no existe. Por favor, verifica tu información.")
            return render(request, "reset_password.html", {"user_id": user_id})  # Conserva el `user_id` original

        # Validar contraseñas: coincidencia y complejidad
        errores = []

        if nueva_contrasena != confirmar_contrasena:
            errores.append("Las contraseñas no coinciden.")

        if nueva_contrasena == confirmar_contrasena:
            # Verificar las condiciones de la contraseña
            if (
                len(nueva_contrasena) < 6 or  # Longitud mínima de 6 caracteres
                not any(c.isalpha() for c in nueva_contrasena) or  # Al menos una letra
                not any(c.isdigit() for c in nueva_contrasena) or  # Al menos un número
                not any(c in "!@#$%^&*()_+-=[]{}|;:',.<>?/`~" for c in nueva_contrasena)  # Al menos un caracter especial
            ):
                errores.append(
                    "La contraseña debe tener al menos 6 caracteres, incluyendo letras, números y caracteres especiales."
                )

        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, "reset_password.html", {"user_id": user_id})  # Conserva el `user_id` original

        # Generar un nuevo salt y encriptar la nueva contraseña
        salt = os.urandom(8).hex()
        nueva_contrasena_encriptada = hashlib.sha256((nueva_contrasena + salt).encode("utf-8")).hexdigest()

        # Actualizar la contraseña y desbloquear la cuenta si estaba bloqueada
        usuario_obj.contrasena = nueva_contrasena_encriptada
        usuario_obj.salt = salt
        if hasattr(usuario_obj, "intentos_fallidos"):
            if usuario_obj.bloqueado:
                usuario_obj.intentos_fallidos = 0  # Reiniciar intentos fallidos
                usuario_obj.bloqueado = False  # Desbloquear la cuenta
                messages.success(request, "Tu cuenta ha sido desbloqueada.")
                messages.success(request, "Dirígete nuevamente a la pagina y inicia sesión con tu nueva contraseña")


        usuario_obj.save()

        messages.success(request, "Tu contraseña ha sido actualizada con éxito ")
        messages.success(request, "Dirígete nuevamente a la pagina y inicia sesión con tu nueva contraseña")


        # return redirect("catalogo") debe mantenerse en la misma página para evitar errores con nombres incorrectos cuando bodeguero o vendedor cambian contraseña
        return render(request, "reset_password.html", {"user_id": user_id})

    # Renderizar la plantilla con el ID en el contexto
    return render(request, "reset_password.html", {"user_id": user_id})


#--------FUNCIONALIDADES CARRITO----------------
def carrito(request):
    try:
        cart = request.session.get('cart', [])
        cart_items = []
        total = 0

        for item in cart:
            producto = Producto.objects.get(id=item['id'])
            subtotal = producto.precio * item['quantity']
            total += subtotal

            cart_items.append({
                'id': producto.id,
                'name': producto.nombre,
                'price': producto.precio,
                'quantity': item['quantity'],
                'subtotal': subtotal,
                'image': producto.img_url
            })

        return render(request, 'carrito.html', {'cart_items': cart_items, 'total': total})
    except Exception as e:
        print(f"Error al cargar el carrito: {e}")
        return render(request, 'carrito.html', {'cart_items': [], 'total': 0})

    
def agregar_producto_carrito(request, producto_id):
    cantidad = request.POST.get('cantidad', 1)
    producto = Producto.objects.get(id=producto_id)

    # Obtener el carrito de la sesión
    carrito = request.session.get('cart', [])

    # Verificar si el producto ya está en el carrito
    producto_en_carrito = next((item for item in carrito if item['id'] == producto.id), None)

    if producto_en_carrito:
        producto_en_carrito['quantity'] += int(cantidad)
    else:
        carrito.append({
            'id': producto.id,
            'quantity': int(cantidad)
        })

    # Guardar el carrito en la sesión
    request.session['cart'] = carrito
    return redirect('carrito')

def obtener_carrito(request):
    carrito = request.session.get('carrito', [])
    return JsonResponse({'cart_items': carrito})
#--------FUNCIONALIDADES CARRITO----------------

def vista_catalogo(request):

    usuario_autenticado = request.session.get('email') is not None  # Verificar si el usuario está autenticado

    query = request.GET.get('search', '')  # Obtener el parámetro 'search' desde la URL
    if query:
        productos = Producto.objects.filter(nombre__icontains=query, estado='habilitado')  # Filtrar por nombre y estado habilitado
        return render(request, 'catalogo.html', {
        'productos': productos, 
        'query': query,
        'usuario_autenticado': usuario_autenticado,  # Agregar al contexto
    })
    else:
        productos = Producto.objects.filter(estado='habilitado')  # Si no hay búsqueda, mostrar solo productos habilitados
        return render(request, 'catalogo.html', {
        'productos': productos, 
        'usuario_autenticado': usuario_autenticado,  # Agregar al contexto
    })



def vista_detalleproducto(request, id):
    # Obtén el producto con el id proporcionado en la URL
    producto = get_object_or_404(Producto, id=id)
    
    # Pasa el producto a la plantilla
    return render(request, 'detalleProducto.html', {'producto': producto})

def cerrar_sesion(request):
    # Limpia la sesión
    if 'cart' in request.session:
       del request.session['cart']  # Elimina el contenido del carrito
    request.session.flush()
    # Crea un mensaje de éxito
    messages.success(request, "Has cerrado sesión exitosamente.")
    return render(request, 'catalogo.html')


def vista_registrouser(request):
    return render(request,'registroUser.html')

def vista_perfiluser(request):
    # Verificar si el usuario está autenticado
    if not request.session.get('email'):
        messages.error(request, "Debes iniciar sesión para acceder a tu perfil.")
        return render(request, 'inicioSesionUser.html')

    # Obtener el usuario autenticado usando el email de la sesión
    email = request.session['email']
    usuario = Usuario.objects.get(email=email)

    if request.method == 'POST':
        # Actualizar los campos con los datos enviados desde el formulario
        usuario.nombre = request.POST.get('nombre', usuario.nombre)
        usuario.primer_apellido = request.POST.get('primer_apellido', usuario.primer_apellido)
        usuario.segundo_apellido = request.POST.get('segundo_apellido', usuario.segundo_apellido)
        usuario.email = request.POST.get('email', usuario.email)
        usuario.telefono = request.POST.get('telefono', usuario.telefono)
        usuario.direccion_particular = request.POST.get('direccion_particular', usuario.direccion_particular)
        usuario.direccion_facturacion = request.POST.get('direccion_facturacion', usuario.direccion_facturacion)

        # Guardar los cambios en la base de datos
        usuario.save()
        messages.success(request, "Tus datos se han actualizado exitosamente.")
        print({usuario.nombre})
        print({usuario.primer_apellido})

        return redirect('perfiluser')

    # Renderizar el formulario con los datos actuales del usuario
    return render(request, 'perfilUser.html', {'usuario': usuario})
def vista_historialuser(request):
    return render(request,'historialCompraUser.html')

# ---------------Admin---------------- 

def vista_inicioadmin(request):
    if request.method == 'POST':
        rut = request.POST.get('rut')
        contrasena = request.POST.get('password')

        # Validación del formato del RUT
        if not validar_rut(rut):
            messages.error(request, 'El RUT ingresado no es válido.')
        try:
            # Consulta para encontrar al administrador por RUT
            admin = Administrador.objects.get(rut=rut)

            # Comparación directa de la contraseña
            if contrasena == admin.contrasena:
                return render(request, 'dashboard.html')
            else:
                messages.error(request, 'La contraseña ingresada es incorrecta.')
        except Administrador.DoesNotExist:
            messages.error(request, 'No está registrado.')

    return render(request, 'inicioAdmin.html')

def validar_rut(rut):
    """
    Valida el formato del RUT chileno. Ejemplo válido: 21270263-3
    """
    import re
    pattern = r'^\d{1,8}-[0-9kK]$'
    return re.match(pattern, rut) is not None

def vista_reporteadmin(request):
    return render(request,'generarReportesAdmin.html')

@role_required(['admin', 'bodeguero', 'vendedor'])
def vista_inventario(request):
    """
    Vista para listar productos en el inventario con soporte para búsqueda.
    """

    # Mostrar todos los productos
    productos = Producto.objects.all()
    print(productos)
    
    query = request.GET.get('search', '')  # Capturar término de búsqueda
    if query:
        # Filtrar productos por nombre o descripción
        productos = Producto.objects.filter(
            Q(nombre__icontains=query) | Q(descripcion__icontains=query)
        )
    else:
        # Mostrar todos los productos
        productos = Producto.objects.all()
        print(productos)

    return render(request, 'verinventario.html', {
        'productos': productos,
        'query': query,
    })

from django.db.models import Q

from django.shortcuts import render, redirect
from .models import Producto

def vista_deshabilitarproducto(request):
    if request.method == 'POST':
        # Recuperar los productos seleccionados y sus nuevos estados
        producto_ids = request.POST.getlist('producto_ids')  # Lista de IDs de los productos seleccionados
        nuevos_estados = request.POST  # Los nuevos estados para cada producto

        for producto_id in producto_ids:
            producto = Producto.objects.get(id=producto_id)
            nuevo_estado = nuevos_estados.get(f'estado_producto_{producto_id}')
            if nuevo_estado:
                producto.estado = nuevo_estado
                producto.save()

        return redirect('deshabilitarproducto')  # Redirigir para evitar reenvíos del formulario

    query = request.GET.get('search', '')  # Recuperar búsqueda si existe
    if query:
        productos = Producto.objects.filter(
            Q(nombre__icontains=query) | Q(id__icontains=query)
        )
    else:
        productos = Producto.objects.all()  # Obtener todos los productos

    return render(request, 'deshabilitarProductoAdmin.html', {
        'productos': productos
    })




def vista_actualizarstock(request):
    # Si el método es POST, actualizamos el stock de los productos
    if request.method == 'POST':
        for producto_id, nueva_cantidad in request.POST.items():
            if producto_id.startswith("producto_"):
                producto_id = producto_id.split("_")[1]
                producto = Producto.objects.get(id=producto_id)
                producto.stock = nueva_cantidad
                producto.save()
                return render(request, 'actualizarStock.html')

    # Si el método es GET, obtenemos los productos y aplicamos el filtro de búsqueda
    else:
        query = request.GET.get('query', '')  # Obtener el valor de búsqueda desde la query string
        if query:
            # Filtrar productos por nombre o ID (asegúrate de ajustar los campos según tus necesidades)
            productos = Producto.objects.filter(
                Q(nombre__icontains=query) | Q(id__icontains=query)
            )
        else:
            productos = Producto.objects.all()  # Si no hay búsqueda, mostrar todos los productos
        
        return render(request, 'actualizarStock.html', {'productos': productos})







from django.shortcuts import render, redirect
from .models import Producto, Promocion
from django.contrib import messages

def vista_panelpromociones(request):
    promociones = Promocion.objects.all()

    if request.method == 'POST':
        for promocion in promociones:
            # Capturar el descuento y el estado desde el formulario
            descuento = request.POST.get(f'descuento_{promocion.id}', 0)
            estado = request.POST.get(f'estado_{promocion.id}', promocion.estado)  # Captura el estado específico de la promoción

            try:
                # Actualizar la promoción en la base de datos
                promocion.descuento = int(descuento)
                promocion.estado = estado
                promocion.save()
                messages.success(request, f"Promoción {promocion.id} actualizada con éxito.")
            except Exception as e:
                messages.error(request, f"Error al actualizar la promoción {promocion.id}: {str(e)}")

        return redirect('panelpromociones')

    return render(request, 'crearPromocionAdmin.html', {'promociones': promociones})






# from django.shortcuts import render, redirect
# from .models import Producto, Promocion
# from django.contrib import messages

# def vista_panelpromociones(request):
#     promociones = Promocion.objects.all()

#     if request.method == 'POST':
#         for promocion in promociones:
#             # Capturar el descuento desde el formulario
#             descuento = request.POST.get(f'descuento_{promocion.id}', 0)
#             estado = request.POST.get('estado', promocion.estado)

#             try:
#                 # Actualizar la promoción en la base de datos
#                 promocion.descuento = int(descuento)
#                 promocion.estado = estado
#                 promocion.save()
#                 messages.success(request, f"Promoción {promocion.id} actualizada con éxito.")
#             except Exception as e:
#                 messages.error(request, f"Error al actualizar la promoción {promocion.id}: {str(e)}")

#         return redirect('panelpromociones')

#     return render(request, 'crearPromocionAdmin.html', {'promociones': promociones})







# from django.shortcuts import render, redirect
# from django.http import JsonResponse
# from .models import Producto, Promocion

# def vista_panelpromociones(request):
#     promociones = Promocion.objects.all()
    
#     # Itera sobre cada promoción en el QuerySet
#     for promocion in promociones:
#         promocion.preciodescuento = int(promocion.preciodescuento)  # Convierte preciodescuento a entero

#     return render(request, 'crearPromocionAdmin.html', {'promociones': promociones})



# def vista_panelpromociones(request):
#     promociones = Promocion.objects.all()
#     promociones.preciodescuento = int(promociones.preciodescuento)
#     return render(request, 'crearPromocionAdmin.html', {'promociones': promociones})

















# @admin_required
def vista_dashboard(request):
    return render(request,'dashboard.html')

@admin_required
def vista_perfiladmin(request):
    return render(request,'perfiladmin.html')





# ---------------vendedor----------------

def vista_perfilvendedor(request):
    return render(request,'perfilvendedor.html') 


def vista_ver_buscar_vendedor(request):
    return render(request,'ver_buscar_vendedor.html')

def vista_panelvendedor(request):
    return render(request,'panelvendedor.html')


# ---------------bodeguero---------------- 
def vista_perfilbodeguero(request):
    return render(request,'perfilbodeguero.html') 


def vista_panelbodeguero(request):
    return render(request,'panelBodeguero.html')


#--------------vistas_añadidas_extras_---------------



def finalizar_compra(request):
    try:
        cart = request.session.get('cart', [])
        cart_items = []
        total = 0

        for item in cart:
            producto = Producto.objects.get(id=item['id'])
            subtotal = producto.precio * item['quantity']
            total += subtotal

            cart_items.append({
                'id': producto.id,
                'name': producto.nombre,
                'price': producto.precio,
                'quantity': item['quantity'],
                'subtotal': subtotal,
                'image': producto.img_url
            })

        return render(request, 'resumen_compra.html', {'cart_items': cart_items, 'total': total})
    except Exception as e:
        print(f"Error al cargar el resumen de la compra: {e}")
        return render(request, 'resumen_compra.html', {'cart_items': [], 'total': 0})



#-------------- Registrar Usuario ---------------------#
def crear_administrador():
    print("=== Crear Administrador ===")
    
    rut = input("Ingrese el RUT del administrador (formato: 7515612-K): ").strip()
    email = input("Ingrese el correo electrónico del administrador: ").strip()
    contrasena = input("Ingrese la contraseña: ").strip()

    # Generar salt y hash de la contraseña
    salt = generar_salt()
    hashed_password = encriptar_con_salt(contrasena, salt)

    # Crear administrador y guardar en la base de datos
    admin = Administrador(
        rut=rut,
        email=email,
        contrasena=hashed_password,
        salt=salt
    )
    admin.save()

    print(f"Administrador creado exitosamente:\nRUT: {rut}\nEmail: {email}")
    print(f"Salt generado: {salt} (se guarda automáticamente en la base de datos)")




from .models import Bodeguero, Producto , Usuario,Administrador, Vendedor 
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.contrib.auth.tokens import default_token_generator
from .models import Vendedor, Bodeguero, Administrador  # Ajusta según tus modelos





# Generar un salt aleatorio
def generar_salt():
    return os.urandom(8).hex()


# Encriptar la contraseña usando el salt
def encriptar_con_salt(password, salt):
    password_bytes = (password + salt).encode('utf-8')
    return hashlib.sha256(password_bytes).hexdigest()


# def validar_campo_unico(request):


#     campo = request.GET.get("campo")
#     valor = request.GET.get("valor")

#     # Verificación para cada campo único
#     if campo == "rut" and Usuario.objects.filter(rut=valor).exists():
#         return JsonResponse({"existe": True, "mensaje": "Este RUT ya está registrado."})
    
#     elif campo == "email" and Usuario.objects.filter(email=valor).exists():
#         return JsonResponse({"existe": True, "mensaje": "Este correo electrónico ya está registrado."})
    


#     elif campo == "telefono" and Usuario.objects.filter(telefono=valor).exists():
#         return JsonResponse({"existe": True, "mensaje": "Este número de teléfono ya está registrado."})

#     return JsonResponse({"existe": False})




def validar_campo_unico(request):
    # Obtener el campo y el valor enviados en la solicitud
    campo = request.GET.get("campo")
    valor = request.GET.get("valor")

    # Validación del campo "rut" en las cuatro tablas
    if campo == "rut":
        if (Vendedor.objects.filter(rut=valor).exists() or 
            Bodeguero.objects.filter(rut=valor).exists() or 
            Usuario.objects.filter(rut=valor).exists() or 
            Administrador.objects.filter(rut=valor).exists()):
            return JsonResponse({"existe": True, "mensaje": "Este RUT ya está registrado."})

    # Validación del campo "email" en las cuatro tablas
    elif campo == "email":
        if (Vendedor.objects.filter(email=valor).exists() or 
            Bodeguero.objects.filter(email=valor).exists() or 
            Usuario.objects.filter(email=valor).exists() or 
            Administrador.objects.filter(email=valor).exists()):
            return JsonResponse({"existe": True, "mensaje": "Este correo electrónico ya está registrado."})

    # Validación del campo "telefono" solo en la tabla Usuario (manteniendo tu lógica original)
    elif campo == "telefono" and Usuario.objects.filter(telefono=valor).exists():
        return JsonResponse({"existe": True, "mensaje": "Este número de teléfono ya está registrado."})

    # Si el valor no existe en ninguna tabla, retornar que no existe
    return JsonResponse({"existe": False})








def activar_cuenta(request, uidb64, token):

    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        usuario = Usuario.objects.get(id=uid)
    except (Usuario.DoesNotExist, ValueError, TypeError):
        usuario = None

    if usuario and default_token_generator.check_token(usuario, token):
        usuario.is_active = True
        usuario.save()
        messages.success(request, "¡Tu cuenta ha sido activada con éxito! Ahora puedes iniciar sesión.")
        return render(request, 'inicioSesionUser.html')
    else:
        messages.error(request, "El enlace de activación no es válido o ha expirado.")


def registrar_usuario(request):
    if request.method == "POST":
        form = UsuarioForm(request.POST)
        if form.is_valid():
            
            usuario = form.save(commit=False)

            # Generar un salt único
            salt = generar_salt()
            usuario.salt = salt

            # Encriptar la contraseña con el salt
            usuario.contrasena = encriptar_con_salt(form.cleaned_data["contrasena"], salt)

            # Poner el usuario como inactivo hasta que confirme
            usuario.is_active = False
            usuario.save()

            # Generar el token de activación
            uid = urlsafe_base64_encode(force_bytes(usuario.id))
            token = default_token_generator.make_token(usuario)

            # Crear enlace de activación
            activation_link = request.build_absolute_uri(
                reverse('activar_cuenta', kwargs={'uidb64': uid, 'token': token})
            )

            # Enviar el correo de activación
            email_subject = "Confirma tu cuenta en El Rincón de los Vecinos"
            email_body = render_to_string('email_confirmacion.html', {
                'nombre_usuario': usuario.nombre,
                'activation_link': activation_link
            })
            email = EmailMultiAlternatives(
                subject=email_subject,
                body=email_body,
                from_email="tuemail@gmail.com",  # Cambia por tu dirección
                to=[usuario.email],
            )
            email.attach_alternative(email_body, "text/html")
            email.send()

            messages.success(request, "Te hemos enviado un correo para confirmar tu cuenta. Revisa tu bandeja de entrada.")
            return render(request, 'catalogo.html')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
            messages.error(request, "Error en los datos ingresados. Revisa los campos.")
    else:
        form = UsuarioForm()
    return render(request, "registroUser.html", {"form": form})






def vista_registroEmpleado(request):
    if request.method == 'POST':

        admin_id = request.session.get('admin_id')

        if not admin_id:
            messages.error(request, "No se pudo identificar al administrador. Por favor, inicie sesión nuevamente.")
            return redirect('inicioSesionUser')

        try:
            # Recuperar la instancia del administrador
            administrador = Administrador.objects.get(id=admin_id)
        except Administrador.DoesNotExist:
            messages.error(request, "El administrador no existe.")
            return redirect('registroEmpleado')
        
        rol = request.POST.get('rolEmpleado', '').strip()
        rut = request.POST.get('rutEmpleado', '').strip()
        email = request.POST.get('emailEmpleado', '').strip()
        contrasena= request.POST.get('passEmpleado', '').strip()
        confirm_contrasena = request.POST.get('confirmPassEmpleado', '').strip()
        
        if not rol:
            messages.error(request, 'Debe seleccionar un tipo de cuenta (Vendedor o Bodeguero).')
            return redirect('registroEmpleado')  # Redirige a la misma página


        if not rol or not rut or not email or not contrasena or not confirm_contrasena:
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('registroEmpleado')
        

        if "@" not in email or "." not in email.split("@")[-1]:
            messages.error(request, "Ingrese un correo electrónico válido.")
            return redirect('registroEmpleado')

        if contrasena != confirm_contrasena:
            messages.error(request, "Las contraseñas no coinciden.")
            return redirect('registroEmpleado')
        
        if contrasena == confirm_contrasena:
            # Verificar las condiciones de la contraseña
            if (
                len(contrasena) < 6 or  # Longitud mínima de 6 caracteres
                not any(c.isalpha() for c in contrasena) or  # Al menos una letra
                not any(c.isdigit() for c in contrasena) or  # Al menos un número
                not any(c in "!@#$%^&*()_+-=[]{}|;:',.<>?/`~" for c in contrasena)  # Al menos un caracter especial

            ):
                messages.error(request, "La contraseña debe tener al menos 6 caracteres, incluyendo letras, números y caracteres especiales.")


        # Validación para evitar duplicados en cualquiera de las tablas
        if Vendedor.objects.filter(rut=rut).exists() or Bodeguero.objects.filter(rut=rut).exists() or Usuario.objects.filter(rut=rut).exists() or Administrador.objects.filter(rut=rut).exists():
            messages.error(request, "El RUT ingresado ya está en uso.")
            return redirect('registroEmpleado')

        if Vendedor.objects.filter(email=email).exists() or Bodeguero.objects.filter(email=email).exists() or Usuario.objects.filter(email=email).exists() or Administrador.objects.filter(email=email).exists():
            messages.error(request, "El correo electrónico ingresado ya está en uso.")
            return redirect('registroEmpleado')
        

        try:
            salt = generar_salt()
            contrasena_encriptada = encriptar_con_salt(contrasena, salt)

            if rol == 'vendedor':
                vendedor = Vendedor(
                    rut=rut,
                    email=email,
                    contrasena=contrasena_encriptada,
                    salt=salt,
                    estado=True,
                    admin_id=administrador,  # Asignar la instancia de Administrador
                )
                vendedor.save()
              
            elif rol == 'bodeguero':
                bodeguero = Bodeguero(
                    rut=rut,
                    email=email,
                    contrasena=contrasena_encriptada,
                    salt=salt,
                    estado=True,
                    admin_id=administrador,  # Asignar la instancia de Administrador
                    
                )
                bodeguero.save()

            messages.success(request, "Empelado registrado con exito")
            return redirect('registroEmpleado')
            

        except Exception as e:
            import traceback
            print(traceback.format_exc())  # Mostrar el error en la consola
            messages.error(request, f"Ocurrió un error: {e}")
            return redirect('registroEmpleado')

    return render(request, 'RegistroVendedor_Bodeguero.html')




import re
from itertools import chain
from django.shortcuts import render, redirect
from django.db.models import Q
from .models import Vendedor, Bodeguero

def validar_rut(rut):
    # Expresión regular para validar el formato del RUT chileno
    rut_regex = r'^\d{7,8}-[\dkK]$'
    return bool(re.match(rut_regex, rut))

def validar_email(email):
    # Validar si el email tiene el formato adecuado
    return email.endswith('@gmail.com')


from django.contrib import messages  # Asegúrate de importar messages

def vista_actualizainfoempleado(request):
    if request.method == 'POST':
        persona_ids = request.POST.getlist('persona_ids')
        nuevos_datos = request.POST

        for persona_id in persona_ids:
            try:
                persona = Vendedor.objects.get(id=persona_id)
                cargo_actual = "vendedor"
            except Vendedor.DoesNotExist:
                persona = Bodeguero.objects.get(id=persona_id)
                cargo_actual = "bodeguero"

            # Obtener datos sensibles
            contrasena = persona.contrasena
            salt = persona.salt
            admin_id = persona.admin_id

            # Actualizar estado
            nuevo_estado = nuevos_datos.get(f'estado_persona_{persona_id}')
            if nuevo_estado:
                persona.estado = nuevo_estado == "activo"

            # Validar y actualizar RUT
            nuevo_rut = nuevos_datos.get(f'rut_persona_{persona_id}')
            if nuevo_rut and not validar_rut(nuevo_rut):
                messages.error(request, f"El RUT {nuevo_rut} no es válido.")
                return render(request, 'actualizainfoempleado.html', {'personas': cargar_personas()})
            if nuevo_rut:
                persona.rut = nuevo_rut


            # Validar y actualizar Email
            nuevo_email = nuevos_datos.get(f'email_persona_{persona_id}')
            if nuevo_email and not validar_email(nuevo_email):
                messages.error(request, f"El correo {nuevo_email} debe ser un correo de Gmail (@gmail.com).")
                return render(request, 'actualizainfoempleado.html', {'personas': cargar_personas()})
            if nuevo_email:
                persona.email = nuevo_email


            # Verificar cambio de cargo
            nuevo_cargo = nuevos_datos.get(f'cargo_persona_{persona_id}')
            if nuevo_cargo and nuevo_cargo != cargo_actual:
                persona.delete()

                if nuevo_cargo == "vendedor":
                    Vendedor.objects.create(
                        rut=nuevo_rut,
                        email=nuevo_email,
                        contrasena=contrasena,
                        salt=salt,
                        estado=(nuevo_estado == "activo"),
                        admin_id=admin_id
                    )
                elif nuevo_cargo == "bodeguero":
                    Bodeguero.objects.create(
                        rut=nuevo_rut,
                        email=nuevo_email,
                        contrasena=contrasena,
                        salt=salt,
                        estado=(nuevo_estado == "activo"),
                        admin_id=admin_id
                    )
                messages.success(request, "Se aplicaron exitosamente los cambios.")
            else:
                persona.save()
                messages.success(request, "Se aplicaron exitosamente los cambios.")

        return redirect('actualizainfoempleado')

    query = request.GET.get('search', '').strip().lower()
    personas = cargar_personas(query)
    return render(request, 'actualizainfoempleado.html', {'personas': personas})


# Nueva función para cargar personas
def cargar_personas(query=''):
    from itertools import chain
    estado_mapeo = {"activo": True, "inactivo": False}

    if query:
        if query in ["vendedor", "bodeguero"]:
            if query == "vendedor":
                return Vendedor.objects.all()
            elif query == "bodeguero":
                return Bodeguero.objects.all()
        else:
            vendedores = Vendedor.objects.filter(
                Q(rut__icontains=query) |
                Q(email__icontains=query) |
                Q(estado=estado_mapeo.get(query, None))
            )
            bodegueros = Bodeguero.objects.filter(
                Q(rut__icontains=query) |
                Q(email__icontains=query) |
                Q(estado=estado_mapeo.get(query, None))
            )
            return list(chain(vendedores, bodegueros))
    else:
        vendedores = Vendedor.objects.all()
        bodegueros = Bodeguero.objects.all()
        return list(chain(vendedores, bodegueros))






